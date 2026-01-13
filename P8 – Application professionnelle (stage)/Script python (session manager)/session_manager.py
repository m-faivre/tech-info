import json
import sys
import os
import pandas as pd
import zipfile
import openpyxl
from functools import partial
from PyQt6.QtWidgets import QApplication, QDialog, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QSpacerItem, QSizePolicy, QLineEdit, QComboBox, QMessageBox, QCheckBox, QWidget, QMainWindow
from PyQt6.QtGui import QFont, QPixmap, QGuiApplication
from PyQt6.QtCore import QTimer, Qt, QPoint
from fiche import FicheWindow
from openpyxl import Workbook
from datetime import datetime
from zipfile import BadZipFile
from openpyxl import load_workbook



# Créer une application Qt
app = QApplication(sys.argv)

# Ne pas quitter l'application lorsque la dernière fenêtre est fermée
app.setQuitOnLastWindowClosed(False)



def load_config():
    try:
        with open('config.json') as f:
            config = json.load(f)
    except FileNotFoundError:
        show_error("Le fichier 'config.json' n'a pas été trouvé dans le répertoire courant.")
        return
    return config



def show_error(message):
    error_dialog = QMessageBox()
    error_dialog.setWindowTitle("Erreur")
    error_dialog.setText(message)
    error_dialog.setIcon(QMessageBox.Icon.Critical)
    error_dialog.exec()
    os._exit(1) 
    
    
global config
config = load_config()
global count_window
count_window = None
    
class FicheEntree(FicheWindow):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.data = {}  # Initialisez self.data dans le constructeur
        self.fields = []
        
        if not os.path.isfile('configuration.exe'):
            show_error("Le fichier 'configuration.exe' n'a pas été trouvé dans le répertoire courant.")

    def show_error(self, message):
        error_dialog = QMessageBox()
        error_dialog.setWindowTitle("Erreur")
        error_dialog.setText(message)
        error_dialog.setIcon(QMessageBox.Icon.Critical)
        error_dialog.exec()
        
    # Boucle les champs présents dans FicheEntreeWindow
    def update_data(self):
        fields = self.findChildren(QLineEdit)
        for field in fields:
            self.data[field.objectName()] = field.text()
        combos = self.findChildren(QComboBox)
        for combo in combos:
            if combo.objectName() == "session_duration_combo":
                self.data["ChoixDuréeSession"] = combo.currentText()
            elif combo.objectName() == "Age":
                self.data["Age"] = combo.currentText()
            elif combo.objectName() == "Statut":
                self.data["Statut"] = combo.currentText()
        if self.venue_box.isChecked():
            self.data["Venue"] = "Oui" 
        else:
            self.data["Venue"] = "Non"
                
    
    # Récupère les valeurs des différents champs
    def get_data(self):
        return self.data
    
    # Ajuste la largeur des colonnes du fichier Excel
    def adjust_column_widths(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width



    # Enregistre les données dans le fichier Excel
    def save_to_excel(self, data):
        current_year = str(datetime.now().year)
        current_date = datetime.now().strftime('%d/%m/%Y')
        current_time = datetime.now().strftime('%H:%M')
        with open('fiche.json', 'r') as f:
            champs = json.load(f)
        ordered_keys = [champ.get("label_content", "").replace(" ", "") for order, champ in sorted(champs.items(), key=lambda item: int(item[0]))]
        ordered_keys.append("ChoixDuréeSession")
        ordered_keys.append("Age")
        ordered_keys.append("Statut")
        ordered_keys.append("Venue")
        new_data_df = pd.DataFrame(data, columns=ordered_keys, index=[0])
        new_data_df.insert(0, "Date", current_date)
        new_data_df.insert(0, "Heure", current_time)
        print(new_data_df) # Pour débugger
        if not os.path.isfile('data.xlsx'):
            new_data_df.to_excel('data.xlsx', sheet_name=current_year, index=False)
        else:
            book = load_workbook('data.xlsx')
            if current_year in book.sheetnames:
                data_df = pd.read_excel('data.xlsx', sheet_name=current_year, engine='openpyxl')
                print(data_df) # Pour débugger
                data_df = pd.concat([data_df, new_data_df], ignore_index=True)
            else:
                data_df = new_data_df
            # Create a new Excel file with the updated data
            with pd.ExcelWriter('data_temp.xlsx', engine='openpyxl') as writer:
                data_df.to_excel(writer, index=False, sheet_name=current_year)
            # Replace the old file with the new one
            os.remove('data.xlsx')
            os.rename('data_temp.xlsx', 'data.xlsx')
        book = load_workbook("data.xlsx")
        sheet = book[current_year]
        self.adjust_column_widths(sheet)
        book.save("data.xlsx")
    
    
    
    # On déclenche les timers à la fermeture de la fenêtre FicheEntreeWindow
    def closeEvent(self, event):
        super().closeEvent(event)
        timer_duree_session()
        timer_popup_2()
        global count_window
        count_window = CountdownWindow(int(config['timer']['duree_session']) * 60)
        count_window.show()
        if config['fermeture']['fermeture_popup']:
            timer_fermeture()
        # Si le log est activé, enregistrer les données dans le fichier Excel
        if config['fiche']['fiche_log']:
            self.update_data()
            #print(self.data) # Pour débugger
            self.save_to_excel(self.get_data())
                # Vérifier que tous les champs sont remplis
        all_fields_filled = all(line_edit.text() for line_edit in self.findChildren(QLineEdit))

        # Vérifier que la case du règlement est cochée
        reglement_checked = self.reglement.isChecked()

        # Réinitialiser le style de tous les champs
        for line_edit in self.findChildren(QLineEdit):
            line_edit.setStyleSheet("")

        if not all_fields_filled or not reglement_checked:
            # Si tous les champs ne sont pas remplis ou que la case du règlement n'est pas cochée, annuler la fermeture de la fenêtre
            event.ignore()

            # Afficher un message d'erreur
            error_message = QMessageBox()
            error_message.setIcon(QMessageBox.Icon.Warning)
            error_message.setWindowTitle("Erreur")
            error_message.setText("Vous devez remplir tous les champs pour vous connecter à votre session.")
            error_message.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint)
            error_message.exec()

            # Encadrer les champs non remplis en rouge
            for line_edit in self.findChildren(QLineEdit):
                if not line_edit.text():
                    line_edit.setStyleSheet("border: 1px solid red;")
        else:
            # Si tous les champs sont remplis et que la case du règlement est cochée, fermer la fenêtre
            event.accept()


# Variables remplacer les {} dans les textes des popups
variables = {
    "session": int(config['timer']['duree_session']),
    "avertissement": (int(config['timer']['duree_session']) - int(config['timer']['timer_popup_1'])),
    "avertissement2": int(config['timer']['timer_popup_2']),
    "fermeture": int(config['timer']['timer_popup_3']),
}

# Fonction pour créer le popup d'après le fichier de configuration
def creation_popup(texte):
    global popup
    popup = QDialog()
    popup.setFixedSize(int(config['style']['largeur_popup']), int(config['style']['hauteur_popup']))
    popup.setStyleSheet(f"background-color: {config['style']['couleur_fond']};")
    popup.setWindowFlags(Qt.WindowType.FramelessWindowHint)
    popup.setWindowFlags(popup.windowFlags() | Qt.WindowType.WindowStaysOnTopHint)
    layout = QVBoxLayout(popup)

    # Créer un nouveau layout pour le label
    layout_texte = QVBoxLayout()
    layout_texte.setContentsMargins(20, 0, 20, 0)  # Appliquer une marge de 10 pixels

    label = QLabel(texte)
    label.setWordWrap(True)
    label.setStyleSheet(f"color : {config['style']['couleur_texte']};")
    police = config['style']['police']
    taille_police = int(config['style']['taille_police'])
    font = QFont(police, taille_police)
    label.setFont(font)

    # Ajouter le label au layout_texte au lieu du layout principal
    layout_texte.addWidget(label)

    # Ajouter le layout_texte au layout principal
    layout.addLayout(layout_texte)

    bouton = QPushButton(f"{config['style']['texte_bouton']}")
    # bouton.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)  # Faire en sorte que le bouton s'adapte au texte
    bouton.setFixedSize(200,35)
    bouton.setStyleSheet(f"""
    QPushButton:enabled {{
        background-color: {config['style']['couleur_bouton']};
        border-radius: 1px;
        border: 1px solid black;
        color: {config['style']['couleur_bouton_texte']};
        font-family: {config['style']['police']};
        font-size: {taille_police}px;
        font-weight: bold;
        padding: 5px 5px;
        text-decoration: none;
    }}
    QPushButton:hover {{
        background-color: {config['style']['couleur_bouton_survol']};
    }}
    QPushButton:pressed {{
        position: relative;
        top: 1px;
    }}
    """)
    bouton.clicked.connect(popup.close)
    layout.addWidget(bouton)
    # Créer un QHBoxLayout
    layout_bouton = QHBoxLayout()
    # Ajouter un QSpacerItem à gauche
    layout_bouton.addItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
    # Ajouter le bouton au layout
    layout_bouton.addWidget(bouton)
    # Ajouter un QSpacerItem à droite
    layout_bouton.addItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
    # Ajouter le layout_bouton au layout principal
    layout.addLayout(layout_bouton)
    popup.setLayout(layout)
    QTimer.singleShot(int(config['timer']['delai_fermeture']) * 1000, popup.close)
    popup.show()
    #print("Popup créée.") # Pour débugger


# Fonction d'affichge de la fenêtre de fermeture
def fermeture(texte):
    global fenetre
    # Créer une fenêtre
    fenetre = QMainWindow()
    # Créer un QLabel pour le texte et le logo
    label_texte = QLabel(texte, fenetre)
    char = config['style']['police']
    taille = int(20)
    police = QFont(char, taille) 
    label_texte.setFont(police)
    label_texte.setStyleSheet(f"color: #476e9e;")
    # Ignorer l'événement de fermeture pour empêcher la fermeture de la fenêtre
    fenetre.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.FramelessWindowHint)
    # Récupérer le répertoire courant du script
    repertoire_courant = os.path.dirname(sys.executable)
    # Définir le chemin du logo
    logo = os.path.join(repertoire_courant, 'img', 'logo.png')
    label_logo = QLabel(fenetre)
    label_logo.setPixmap(QPixmap(logo))
    # Centrer le texte
    label_texte.setAlignment(Qt.AlignmentFlag.AlignCenter)
    # Créer un layout vertical et y ajouter le logo et le texte
    layout = QVBoxLayout()
    layout.addWidget(label_logo, alignment=Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
    layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))  # Ajouter un QSpacerItem avant le texte
    layout.addWidget(label_texte)
    layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))  # Ajouter un QSpacerItem après le texte
    # Créer un widget central pour la fenêtre et y appliquer le layout
    widget_central = QWidget()
    widget_central.setLayout(layout)
    fenetre.setCentralWidget(widget_central)
    # Créer un QTimer pour fermer la fenêtre après le délai
    timer = QTimer(app)
    timer.timeout.connect(fenetre.close)
    timer.start(int(config['timer']['timer_popup_3']) * 1000)  # Convertir le délai en millisecondes
    #timer.start(int(5) * 1000)  # DEBUG
    # Mettre la fenêtre en plein écran
    fenetre.showFullScreen()
    # Afficher la fenêtre
    fenetre.show()
    # app.exec()



# Fonction pour créer le timer de la durée de la session
def timer_duree_session():
    # Convertir la durée de la session en secondes
    duree_session = (int(config['timer']['duree_session']) * 60) + int(config['timer']['timer_popup_3'])
    #print(f"Durée session : {duree_session}") # Pour débugger
    # Créer un timer
    global timer
    timer = QTimer(app)
    # Connecter le signal timeout du timer à la fonction de fermeture de session
    timer.timeout.connect(end_session)
    # Démarrer le timer
    timer.start(duree_session * 1000)  # Convertir les secondes en millisecondes
    # Imprimer un message pour indiquer que le timer est actif
    #print("Le timer durée session est actif.") # Pour débugger
    timer_popup_1()



# Fonction pour créer le timer du premier popup
def timer_popup_1():
    timer_1 = (int(config['timer']['timer_popup_1']) * 60)
    global timer1
    #print(f"Timer popup 1 : {timer_1}") # Pour débugger
    timer1 = QTimer(app)
    timer1.stop()  # Arrêter le timer avant de le configurer
    texte = config['text']['text_popup_1'].format(**variables)
    creation_popup_1 = partial(creation_popup, texte)
    timer1.timeout.connect(creation_popup_1)
    timer1.timeout.connect(timer1.stop)
    timer1.start(timer_1 * 1000) # DEBUG
    #print("Le timer popup 1 est actif.") # Pour débugger



# Fonction pour créer le timer du deuxième popup
def timer_popup_2():
    global timer_2
    timer_2 = (int(config['timer']['duree_session']) * 60) - (int(config['timer']['timer_popup_2']) * 60)
    #print(f"Timer popup 2 : {timer_2}") # Pour débugger
    global timer2
    timer2 = QTimer(app)
    timer2.stop()  # Arrêter le timer avant de le configurer
    texte = config['text']['text_popup_2'].format(**variables)
    creation_popup_2 = partial(creation_popup, texte)
    timer2.timeout.connect(creation_popup_2)
    timer2.timeout.connect(timer2.stop)
    timer2.start(timer_2 * 1000) # Convertir les secondes en millisecondes
    #print("Le timer popup 2 est actif.") # Pour débugger
    
    

# Fonction pour créer le timer de fermeture de la session
def timer_fermeture():
    timerfermeture = (int(config['timer']['duree_session']) * 60)
    #print(f"Timer fermeture : {timerfermeture}") # Pour débugger
    global timerfin
    timerfin = QTimer()
    timerfin.stop()  # Arrêter le timer avant de le configurer
    texte = config['text']['text_fermeture'].format(**variables)
    creation_fin = partial(fermeture, texte)
    timerfin.timeout.connect(creation_fin)
    timerfin.timeout.connect(timerfin.stop)
    timerfin.start(timerfermeture * 1000)
    
    

# Fonction appelée par duree_session : déconnecte ou verrouille la session
def end_session():
    # Fermer la session en fonction de la valeur de session_fermeture
    if config['fermeture']['fermeture_session'] == 'Déconnecter':
        # Déconnecter l'utilisateur
        os.system('logoff')
        #print("Fin de session. Déconnexion.") # Pour débugger
    else:
        # Verrouiller la session
        os.system('rundll32.exe user32.dll,LockWorkStation')
        #print("Fin de session. Verrouillage.") # Pour débugger
    sys.exit(app.exec()) # Terminer l'application Qt



class CountdownWindow(QWidget):
    def __init__(self, timeout, parent=None):
        super(CountdownWindow, self).__init__(parent)
        self.time_left = timeout

        self.label = QLabel(self.format_time(self.time_left), self)
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setFont(QFont(config['style']['police'], int(config['style']['taille_police'])))  # Définit la police et la taille du texte
        self.label.setStyleSheet(f"color: {config['style']['couleur_texte']};")  # Définit la couleur du texte

        layout = QVBoxLayout(self)
        layout.addWidget(self.label)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_countdown)
        self.timer.start(1000)

        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.oldPos = self.pos()

        self.setFixedSize(200, 100)  # Définit la taille de la fenêtre

        screen_geometry = QGuiApplication.primaryScreen().geometry()
        x = (screen_geometry.width() - self.width()) // 2
        self.move(x, 0)

    def format_time(self, seconds):
        minutes, seconds = divmod(seconds, 60)
        return f"Temps restant :\n {minutes:02d}:{seconds:02d}"

    def update_countdown(self):
        self.time_left -= 1
        if self.time_left <= 0:
            self.timer.stop()
            self.label.setText("Terminé !")
            self.close()
        else:
            self.label.setText(self.format_time(self.time_left))

    def closeEvent(self, event):
        if self.time_left > 0:
            event.ignore()

    # def mousePressEvent(self, event):
    #     self.oldPos = event.globalPosition()

    # def mouseMoveEvent(self, event):
    #     delta = QPoint(event.globalPosition().toPoint() - self.oldPos.toPoint())
    #     self.move(self.x() + delta.x(), self.y() + delta.y())
    #     self.oldPos = event.globalPosition()



# Créer et afficher la fenêtre FicheEntreeWindow
window = FicheEntree()
window.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.FramelessWindowHint)
window.showFullScreen()

app.exec()
# sys.exit(app.exec())
